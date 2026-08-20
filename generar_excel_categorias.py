from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Categorias ASORETA"

header_fill = PatternFill(start_color="B90D09", end_color="B90D09", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="B90D09")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.merge_cells("A1:E1")
ws["A1"] = "ASORETA - Análisis de categorías múltiples"
ws["A1"].font = title_font

ws.merge_cells("A3:E3")
ws["A3"] = "CUADRO 1: Cómo están AHORA (1 sola categoría por lugar)"
ws["A3"].font = Font(bold=True, size=12)

headers1 = [
    "Establecimiento",
    "Categoría actual (única)",
    "Observación",
    "Visible en categoría secundaria?",
    "Notas",
]
for col, h in enumerate(headers1, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

current = [
    ["Manakish", "Árabe", 'Nombre incluye "DULCERIA Y LICORES"', "No", "Solo aparece en Árabe"],
    ["La Terraza Café Panadería Y Pastelería C.A", "Internacional", "Nombre incluye Café y Panadería", "No", "Solo aparece en Internacional"],
    ["Primos Café & Rest.", "Internacional", "Nombre incluye Café", "No", "Solo aparece en Internacional"],
    ["Waku Casa Cafe", "Internacional", "Es café + restaurante", "No", "Solo aparece en Internacional"],
    ["Delicious Malteadas", "Cafeterías y panadería", "Vende malteadas/postres", "No", "No aparece en Postres"],
    ["Gin Restobar", "Bar & lounge", "Restobar con cocina variada", "No", "No aparece en Internacional"],
    ["40 Grados", "Bar & lounge", "Bar con comida", "No", "No aparece en Internacional"],
    ["Entre Sabores Cafe", "Cafeterías y panadería", "Café especializado", "No", "Podría ser Internacional"],
    ["Gabys La Mía Pizzeria", "ITALIANA (inconsistente)", "Badge en mayúsculas", "No", 'Corregir a "Italiana"'],
    ["Nico Gelato Gelateria", "Postres, helados y galleteria", "También cafetería en CC", "No", "Podría ser Cafeterías"],
    ["La Cioccolata", "Italiana", "También repostería/chocolate", "No", "Podría ser Postres"],
]

for r, row in enumerate(current, 5):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

start = r + 3
ws.merge_cells(f"A{start}:E{start}")
ws[f"A{start}"] = "CUADRO 2: Propuesta con 2+ categorías (tags múltiples con data-categories)"
ws[f"A{start}"].font = Font(bold=True, size=12)

headers2 = [
    "Establecimiento",
    "Categoría principal (badge visible)",
    "Categorías adicionales",
    "Aparecería en galería de...",
    "Beneficio para el usuario",
]
for col, h in enumerate(headers2, 1):
    cell = ws.cell(row=start + 1, column=col, value=h)
    cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

proposed = [
    ["Manakish", "Árabe", "Postres, helados y galleteria", "Árabe + Postres", "Quien busca dulces también lo encuentra"],
    ["La Terraza Café Panadería Y Pastelería C.A", "Internacional", "Cafeterías y panadería", "Internacional + Cafeterías", "Refleja su nombre real"],
    ["Primos Café & Rest.", "Internacional", "Cafeterías y panadería", "Internacional + Cafeterías", "Café gourmet visible en ambas"],
    ["Waku Casa Cafe", "Internacional", "Cafeterías y panadería", "Internacional + Cafeterías", "Café como entrada principal"],
    ["Delicious Malteadas", "Cafeterías y panadería", "Postres, helados y galleteria", "Cafeterías + Postres", "Malteadas = postres/bebidas"],
    ["Gin Restobar", "Bar & lounge", "Internacional", "Bar & lounge + Internacional", "Restobar con carta variada"],
    ["40 Grados", "Bar & lounge", "Internacional", "Bar & lounge + Internacional", "Bar con oferta gastronómica"],
    ["Entre Sabores Cafe", "Cafeterías y panadería", "Internacional", "Cafeterías + Internacional", "Café de especialidad"],
    ["Gabys La Mía Pizzeria", "Italiana", "(ninguna adicional)", "Solo Italiana", "Unificar mayúsculas → Italiana"],
    ["Nico Gelato Gelateria", "Postres, helados y galleteria", "Cafeterías y panadería", "Postres + Cafeterías", "Gelatería en centro comercial"],
    ["La Cioccolata", "Italiana", "Postres, helados y galleteria", "Italiana + Postres", "Chocolate y repostería italiana"],
]

for i, row in enumerate(proposed, start + 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

for i, w in enumerate([35, 28, 30, 28, 38], 1):
    ws.column_dimensions[chr(64 + i)].width = w

ws2 = wb.create_sheet("Resumen")
ws2["A1"] = "Resumen del directorio"
ws2["A1"].font = title_font
summary = [
    ("Total tarjetas en el sitio", "62"),
    ("Categoría única por lugar (ahora)", "62 de 62"),
    ("Lugares candidatos a 2+ categorías", "12 identificados"),
    ("Implementados con data-categories", "Manakish, La Terraza, Primos, Waku"),
    ("Estadística corregida en sitio", "62 en directorio digital"),
    ("Nota", "Los conteos de galería se actualizan automáticamente con tags múltiples"),
]
for i, (k, v) in enumerate(summary, 3):
    ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws2.cell(row=i, column=2, value=v)
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 50

out = "ASORETA_categorias_multiples.xlsx"
wb.save(out)
print("Saved:", out)
